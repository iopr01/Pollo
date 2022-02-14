import openpyxl

headers = ['Index', ' ', ' ', 'Datos', ' ', ' ', "tiempo"]
headers_2 = [' ','Temperatura', 'Humedad', 'Tanque', 'peso', 'comida', ' ']


def create_load_workbook(filename):
    # Loading workbook and getting first spreadsheet
    # If the file doesn't exists, it is created
    # noinspection PyBroadException
    try:
        workbook = openpyxl.load_workbook(filename=filename)  # load workbook with the name
        sheet = workbook['Temperatura']
    except Exception:
        workbook = openpyxl.Workbook()
        elim = workbook['Sheet']
        workbook.remove(elim)
        sheet = workbook.create_sheet('Temperaturas')
        for i in range(len(headers)):
            sheet.cell(1, i + 1).value = headers[i]
            sheet.cell(1, i + 1).font = openpyxl.styles.Font(bold=True)
            sheet.cell(2, i + 1).value = headers_2[i]
            sheet.cell(2, i + 1).font = openpyxl.styles.Font(bold=True)
    return workbook, sheet


def write(sheet, data):
    blank_row = sheet.max_row
    print(blank_row, data)
    sheet.cell(blank_row, 1).value = blank_row - 1
    for i, dat in enumerate(data):
        offset = 1 if i > 5 else 0
        sheet.cell(blank_row, i+offset+3).value = dat
